from datetime import datetime, timedelta
import pytz
import json
import os
import logging
import pandas as pd
import requests
from dotenv import load_dotenv
from typing import Dict, List, Optional
from requests.exceptions import HTTPError, RequestException
from functools import lru_cache

import aiohttp
import asyncio

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

load_dotenv()

class MercadoLivreAPIError(Exception):
    """Exceção personalizada para erros da API do Mercado Livre"""
    pass

class MLTokenManager:
    """Gerencia autenticação e renovação de tokens do Mercado Livre"""
    
    def __init__(self, client_id: str, client_secret: str):
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token: Optional[str] = None
        self.refresh_token_value: Optional[str] = None  # Atributo renomeado

    def authenticate(self) -> None:
        """Obtém novo token de acesso usando client credentials"""
        try:
            response = requests.post(
                "https://api.mercadolibre.com/oauth/token",
                data={
                    "grant_type": "client_credentials",
                    "client_id": self.client_id,
                    "client_secret": self.client_secret
                },
                timeout=10
            )
            response.raise_for_status()
            token_data = response.json()
            self.access_token = token_data.get("access_token")
            self.refresh_token_value = token_data.get("refresh_token")
            logger.info("Autenticação realizada com sucesso")
            
        except HTTPError as e:
            logger.error(f"Erro de autenticação: {e.response.text}")
            raise MercadoLivreAPIError("Falha na autenticação") from e

    def renew_token(self) -> None:  # Método renomeado
        """Renova o token de acesso usando refresh token"""
        if not self.refresh_token_value:
            raise MercadoLivreAPIError("Refresh token não disponível")
            
        try:
            response = requests.post(
                "https://api.mercadolibre.com/oauth/token",
                data={
                    "grant_type": "refresh_token",
                    "client_id": self.client_id,
                    "client_secret": self.client_secret,
                    "refresh_token": self.refresh_token_value
                },
                timeout=10
            )
            response.raise_for_status()
            token_data = response.json()
            self.access_token = token_data.get("access_token")
            logger.info("Token de acesso renovado com sucesso")
            
        except HTTPError as e:
            logger.error(f"Erro ao renovar token: {e.response.text}")
            raise MercadoLivreAPIError("Falha ao renovar token") from e

class MercadoLivreAPI:
    """Classe principal para integração com a API do Mercado Livre"""
    
    def __init__(self):
        self.client_id = os.getenv("MERCADO_LIVRE_CLIENT_ID")
        self.client_secret = os.getenv("MERCADO_LIVRE_CLIENT_SECRET")
        self.user_id = os.getenv("MERCADO_LIVRE_USER_ID")
        self.token_manager = MLTokenManager(self.client_id, self.client_secret)

        #teste    
        self.GROQ_API_KEY = os.getenv("GROQ_API_KEY")
        
        self._validate_credentials()
        self._setup_session()

    def _validate_credentials(self) -> None:
        """Valida as credenciais essenciais"""
        missing = []
        if not self.client_id:
            missing.append("MERCADO_LIVRE_CLIENT_ID")
        if not self.client_secret:
            missing.append("MERCADO_LIVRE_CLIENT_SECRET")
        if not self.user_id:
            missing.append("MERCADO_LIVRE_USER_ID")
            
        if missing:
            raise MercadoLivreAPIError(
                f"Credenciais faltando no .env: {', '.join(missing)}"
            )

    def _setup_session(self) -> None:
        """Configura sessão HTTP reutilizável"""
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "DVSmartShop/1.0",
            "Accept": "application/json",
            "Content-Type": "application/json"
        })

    def _make_request(self, url: str, params: Optional[Dict] = None) -> Dict:
        """Executa requisições à API com tratamento de erros"""
        try:
            if not self.token_manager.access_token:
                self.token_manager.authenticate()
                
            headers = {"Authorization": f"Bearer {self.token_manager.access_token}"}
            
            response = self.session.get(
                url,
                headers=headers,
                params=params,
                timeout=60
            )
            logger.debug(f"Resposta bruta da API: {response.text}")  # Para debug
            response.raise_for_status()
            
            return response.json()
            
        except HTTPError as e:
            if e.response.status_code == 401:
                logger.warning("Token expirado, tentando renovar...")
                self.token_manager.renew_token()
                return self._make_request(url, params)
            logger.error(f"Erro na API: {e.response.text}")
            raise MercadoLivreAPIError("Erro na requisição à API") from e
            
        except RequestException as e:
            logger.error(f"Erro de conexão: {str(e)}")
            raise MercadoLivreAPIError("Erro de conexão") from e

    @lru_cache(maxsize=128)
    def _get_active_items(self) -> List[str]:
        """
        Obtém IDs de itens ativos com cache, excluindo anúncios de catálogo.
        """
        try:
            # Faz a requisição para obter os IDs dos anúncios ativos
            response = self._make_request(
                f"https://api.mercadolibre.com/users/{self.user_id}/items/search",
                params={"status": "active"}
            )
            
            # Extrai os resultados da resposta
            results = response.get("results", [])
            
            # Filtra os IDs dos anúncios que não são de catálogo
            non_catalog_ids = []
            for item_id in results:
                item_details = self._make_request(f"https://api.mercadolibre.com/items/{item_id}")
                if not item_details.get("catalog_listing", False):  # Verifica se não é um anúncio de catálogo
                    non_catalog_ids.append(item_id)
            
            print(non_catalog_ids)
            return non_catalog_ids

        except MercadoLivreAPIError as e:
            logger.error(f"Falha ao obter itens ativos: {str(e)}")
            return []        

    def _process_item_data(self, item_data: Dict) -> List[Dict]:
        """Processa os dados de um item para extrair estoque e SKU"""
        try:
            sku = next(
                (attr["value_name"] for attr in item_data.get("attributes", [])
                if attr.get("id") == "SELLER_SKU"
            ), None)
            
            variations = item_data.get("variations", [])
            processed = []

            # Processa variações ou item único
            items_to_process = variations if variations else [item_data]
            for item in items_to_process:
                processed.append({
                    "SKU": self._extract_sku(item) or sku,
                    "Nome": item_data.get("title", "")[:70],
                    "Estoque": item.get("available_quantity", 0)
                })
                
            return processed
            
        except Exception as e:
            logger.warning(f"Erro ao processar item {item_data.get('id')}: {str(e)}")
            return []

    def _extract_sku(self, item: Dict) -> Optional[str]:
        """Extrai SKU de um item ou variação"""
        return next(
            (attr["value_name"] for attr in item.get("attributes", [])
             if attr.get("id") == "SELLER_SKU"),
            None
        )

    def gerar_relatorio_estoque(self) -> pd.DataFrame:
        """Gera relatório consolidado de estoque"""
        try:
            logger.info("Obtendo dados de estoque do Mercado Livre...")
            item_ids = self._get_active_items()
            
            stock_data = []
            for item_id in item_ids:
                item_data = self._make_request(
                    f"https://api.mercadolibre.com/items/{item_id}",
                    params={"include_attributes": "all"}
                )
                stock_data.extend(self._process_item_data(item_data))
                
            return self._create_dataframe(stock_data)
            
        except MercadoLivreAPIError as e:
            logger.error(f"Erro crítico: {str(e)}")
            return pd.DataFrame()

    def _create_dataframe(self, raw_data: List[Dict]) -> pd.DataFrame:
        """Cria DataFrame padronizado com tratamento de tipos"""
        try:
            df = pd.DataFrame(raw_data)
            
            # Conversão segura de tipos numéricos
            if 'Estoque' in df.columns:
                df['Estoque'] = pd.to_numeric(df['Estoque'], errors='coerce').fillna(0).astype(int)
            
            # Ordena colunas e remove duplicatas
            df = df[['SKU', 'Nome', 'Estoque']].drop_duplicates()
            
            return df
            
        except Exception as e:
            logger.error(f"Erro na criação do DataFrame: {str(e)}")
            return pd.DataFrame()
        

    async def _make_request_async(self, session, url: str, params: Optional[Dict] = None) -> Dict:
        """Versão assíncrona de _make_request"""
        try:
            if not self.token_manager.access_token:
                await self.token_manager.authenticate_async()
            
            headers = {"Authorization": f"Bearer {self.token_manager.access_token}"}
            
            async with session.get(url, headers=headers, params=params, timeout=15) as response:
                logger.debug(f"Resposta bruta da API: {await response.text()}")  # Para debug
                response.raise_for_status()
                return await response.json()
                
        except aiohttp.ClientResponseError as e:
            if e.status == 401:
                logger.warning("Token expirado, tentando renovar...")
                await self.token_manager.renew_token_async()
                return await self._make_request_async(session, url, params)
            logger.error(f"Erro na API: {e.message}")
            raise MercadoLivreAPIError("Erro na requisição à API") from e
            
        except aiohttp.ClientError as e:
            logger.error(f"Erro de conexão: {str(e)}")
            raise MercadoLivreAPIError("Erro de conexão") from e
        
        
    @lru_cache(maxsize=128)
    async def get_shipment_details_async(self, session, shipping_id: str) -> tuple[str, float]:
        """
        Versão assíncrona de get_shipment_details
        """
        url = f"https://api.mercadolibre.com/shipments/{shipping_id}"
        shipment_details = await self._make_request_async(session, url)
        logistic_type = shipment_details.get('logistic_type')
        logistic_cost = shipment_details.get('shipping_option', {}).get('cost', 0.0)
        return logistic_type, logistic_cost


    async def get_all_shipment_details(self, shipping_ids):
        async with aiohttp.ClientSession() as session:
            tasks = [self.get_shipment_details_async(session, shipping_id) for shipping_id in shipping_ids]
            return await asyncio.gather(*tasks)    



    def get_sales_data(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Recupera dados de vendas do Mercado Livre para o período especificado.
        
        :param start_date: Data de início no formato "dd/mm/yyyy"
        :param end_date: Data de fim no formato "dd/mm/yyyy"
        :return: DataFrame com os dados de vendas
        """
        try:
 
            # Defina as datas em UTC-3 (Brasília)
            start_date_brt = datetime.strptime(start_date, "%d/%m/%Y")
            end_date_brt = datetime.strptime(end_date, "%d/%m/%Y")

            # Converta para UTC-4 (subtraindo 1 hora)
            brt = pytz.timezone('America/Sao_Paulo')
            start_utc4 = brt.localize(start_date_brt).astimezone(pytz.timezone('Etc/GMT+4')) - timedelta(hours=1)
            end_utc4 = brt.localize(end_date_brt).astimezone(pytz.timezone('Etc/GMT+4')) + timedelta(hours=23, minutes=59, seconds=59, microseconds=999999)            

            print(f' Datas meli.py -> Início: {start_utc4} - Fim: {end_utc4}')

            url = "https://api.mercadolibre.com/orders/search"
            # Formate os parâmetros da API
            params = {
                "seller": self.user_id,
                "order.date_closed.from": start_utc4.isoformat(),
                "order.date_closed.to": end_utc4.isoformat(),                
                #"order.status": "paid",
                "limit": 50
            }

            all_orders = []
            while True:
                response = self._make_request(url, params)
                logger.debug(f"Resposta da API: {response}")  # Log da resposta completa
                results = response.get('results', [])
                all_orders.extend(results)

                # Write the response results to a file
                #with open('api_response.txt', 'w', encoding='utf-8') as file:
                #    file.write(json.dumps(results, indent=2, ensure_ascii=False) + '\n')

                if len(results) < params['limit']:
                    break

                params['offset'] = params.get('offset', 0) + params['limit']

            logger.info(f"Total de pedidos recuperados: {len(all_orders)}")

            sales_data = []
            qtd_total = 0
            fat_total = 0
            
            all_shipping_ids = [order['shipping']['id'] for order in all_orders if 'shipping' in order]
            unique_shipping_ids = list(set(all_shipping_ids))  # Remove duplicatas

            # Obtenha os detalhes de envio de forma assíncrona
            shipment_details = asyncio.run(self.get_all_shipment_details(unique_shipping_ids))
            shipment_dict = dict(zip(unique_shipping_ids, shipment_details))

            
            for order in all_orders:
                try:
                    date_created_utc4 = datetime.fromisoformat(order['date_closed'])
                    date_created_brt = date_created_utc4.astimezone(brt)

                    if start_date_brt.date() <= date_created_brt.date() <= end_date_brt.date():
                        

                        for item in order['order_items']:
                            #sku = item['item'].get('seller_custom_field') or item['item']['id']
                            sku = item['item']['seller_sku']
                            qty = item['quantity']
                            unit_price = item['unit_price']
                            product_name = item['item'].get('title')

                            # Extrair informações de pagamento
                            payments = order.get('payments')
                            order_id = payments[0]['order_id']
                            payment_status = payments[0]['status']

                            # Recupera valor efetivo pago pelo usuário
                            paid_amount = order.get('paid_amount')

                            # Recupera o valor do frete do comprador e a malha logística da venda
                            shipping_id = order['shipping']['id']                            
                            logistic_type, logistic_cost = shipment_dict.get(shipping_id, ("unknown", 0.0))
                            #logistic_type, logistic_cost = shipment_details.get(shipping_id, ("unknown", 0.0))

                            # verificar por que está lento
                            #logistic_type, logistic_cost = self.get_shipment_details(shipping_id)
                            

                            order_status = order['status']

                            # com frete comprador
                            paid_amount_calculated = paid_amount + logistic_cost #(unit_price * qty) + logistic_cost

                            # Cenários de reembolso parcial
                            if order_status == 'partially_refunded':
                                # com frete comprador
                                paid_amount_calculated = paid_amount + logistic_cost 

                                # sem frete comprador (default Mercado turbo)
                                paid_amount_calculated_no_ship_cost = paid_amount
                                                                
                            else:    
                                # com frete comprador
                                paid_amount_calculated = (unit_price * qty) + logistic_cost

                                # sem frete comprador (default Mercado turbo)
                                paid_amount_calculated_no_ship_cost = unit_price * qty

                            sales_data.append({
                                'order_id': order_id,
                                'order_status': order_status,
                                'payment_status': payment_status,
                                'product_name': product_name,
                                'date': date_created_brt,
                                'sku': sku,
                                'qty': qty,
                                'unit_price': unit_price,
                                'paid_amount': paid_amount,
                                'paid_amount_calculated_no_ship_cost': paid_amount_calculated_no_ship_cost,   #default mercado turbo - sem frete comprador                            
                                'paid_amount_calculated': paid_amount_calculated,                                  
                                'logistic_cost': logistic_cost,
                                'logistic_type': logistic_type,
                            })

                            

                            qtd_total += qty
                            fat_total += paid_amount_calculated_no_ship_cost
                            
                    else:
                        logger.warning(f"Pedido {order['id']} fora do intervalo de datas: {date_created_brt.date()}")
                    

                except KeyError as ke:
                    logger.warning(f"Chave não encontrada ao processar pedido: {ke}")
                    logger.debug(f"Detalhes do pedido: {order}")

            logger.info(f"Total de registros de venda processados: {len(sales_data)}")            
            logger.info(f"Quantidade total unidades vendidas: {qtd_total}")
            logger.info(f"Faturamento Total: {fat_total}")                          

            return pd.DataFrame(sales_data)

        except Exception as e:
            logger.error(f"Erro ao obter dados de vendas: {str(e)}", exc_info=True)
            return pd.DataFrame()

    def fetch_all_item_ids(self, seller_id: str, limit: int = 50) -> list:
        """
        Recupera todos os item_ids (anúncios) de um vendedor no Mercado Livre.
        """
        url = "https://api.mercadolibre.com/sites/MLB/search"
        offset = 0
        all_item_ids = []

        while True:
            params = {
                "seller_id": seller_id,
                "limit": limit,
                "offset": offset
            }
            try:
                response = self._make_request(url, params)
                results = response.get("results", [])
            except Exception as err:
                logger.error(f"Erro na requisição: {err}")
                break

            if not results:
                break

            all_item_ids.extend([item.get("id") for item in results])

            # Paginação
            if len(results) < limit:
                break
            offset += limit

        return all_item_ids
    
               
    def get_all_questions_answers(self, item_id: str, limit: int = 50) -> list:
        """
        Recupera todas as perguntas e respostas de um anúncio no Mercado Livre.
        """
        url = "https://api.mercadolibre.com/questions/search"
        offset = 0
        all_entries = []

        while True:
            params = {
                "api_version": "4",
                "item": item_id,
                "limit": limit,
                "offset": offset
            }
            try:
                response = self._make_request(url, params)                
                questions = response.get("questions", [])
            except Exception as err:
                logger.error(f"Erro na requisição: {err}")
                break

            if not questions:
                break

            for q in questions:
                entry = {
                    "question_id": q.get("id"),
                    "question_text": q.get("text"),
                    "question_date": q.get("date_created"),
                    "question_status": q.get("status"),
                    "answer_text": q.get("answer", {}).get("text") if q.get("answer") else None,
                    "answer_date": q.get("answer", {}).get("date_created") if q.get("answer") else None,
                    "answer_status": q.get("answer", {}).get("status") if q.get("answer") else None
                }
                all_entries.append(entry)

            # Paginação
            if len(questions) < limit:
                break
            offset += limit

        return all_entries

    def generate_general_report(self, start_date: str, end_date: str) -> str:
        df = self.get_sales_data(start_date, end_date)
        if df.empty:
            return f"Nenhum dado de venda encontrado para o período de {start_date} a {end_date}."
        
        # Truncar a data e formatar para DD/MM/YYYY
        #df['date'] = df['date'].dt.strftime('%d/%m/%Y')

        '''
        try: 
            file_name = f"Relatorio_Modalidade_{start_date.replace('/', '_')}_{end_date.replace('/', '_')}.csv"
            
            df.to_csv(file_name, decimal=',')
            print('CSV de vendas gerado com sucesso!')
        except Exception as e:
            logger.error(f"Erro ao gerar csv de vendas: {str(e)}", exc_info=True)
        '''    


        #Agrupado por mês/ano
        df['date'] = df['date'].dt.strftime('%m/%Y')

        #df_vendas_aprovadas = df[df['order_status'] == 'paid']
        #df_vendas_aprovadas = df.query("order_status != 'cancelled' and sku == 'DVLAVADORAPREMIUM'")
        df_vendas_aprovadas = df.query("order_status != 'cancelled'")
                
        df_grouped = df_vendas_aprovadas.groupby(['date', 'sku']).agg({
            'qty': 'sum',
            'paid_amount_calculated_no_ship_cost': 'sum'
        }).reset_index()
        
        report = f"Relatório Geral\nPeríodo: {start_date} a {end_date}\n\n"
        for _, row in df_grouped.sort_values(['date', 'qty'], ascending=[False, False]).iterrows():
            report += f"{row['date']} - {row['sku']} - {row['qty']} unidades - R$ {row['paid_amount_calculated_no_ship_cost']:.2f}\n"
        
        # Calcular totais
        total_qty = df_vendas_aprovadas['qty'].sum()
        #total_faturamento = df_vendas_aprovadas['paid_amount'].sum()
        total_faturamento = df_vendas_aprovadas['paid_amount_calculated_no_ship_cost'].apply(lambda x: round(x, 2)).sum() 

        print("\n" + "="*60 + "\n")
        
        # Adicionar resumo ao final do relatório
        report += f"\nResumo:\n"
        report += f"Total de itens vendidos: {total_qty}\n"
        report += f"Faturamento total: R$ {total_faturamento:.2f}\n"
        
        return report
    
    def generate_salesdash(self, start_date: str, end_date: str):
        df = self.get_sales_data(start_date, end_date)

        if df.empty:
            return f"Nenhum dado de venda encontrado para o período de {start_date} a {end_date}."                
                
        df_vendas_aprovadas = df.query("order_status != 'cancelled'")                                        
                       
        return df_vendas_aprovadas


    def generate_modality_report(self, start_date: str, end_date: str) -> str:
        df = self.get_sales_data(start_date, end_date)
        if df.empty:
            return f"Nenhum dado de venda encontrado para o período de {start_date} a {end_date}."
                        
        report = f"Relatório por Modalidade\nPeríodo: {start_date} a {end_date}\n\n"

        for _, row in df.sort_values(['date', 'qty'],ascending=[False, False]).iterrows():
            report += f"{row['date']} - {row['order_id']} - {row['sku']} - {row['product_name']} - {row['order_status']} - {row['payment_status']} - {row['qty']} unidades - R$ {row['unit_price']:.2f} - R$ {row['paid_amount_calculated_no_ship_cost']:.2f} - {row['logistic_cost']} - {row['logistic_type']}\n"
            #report += f"#{row['order_id']} \n"
        
        return report 
    
    def generate_modality_report_excel(self, start_date: str, end_date: str) -> str:
        df = self.get_sales_data(start_date, end_date)
        if df.empty:
            return f"Nenhum dado de venda encontrado para o período de {start_date} a {end_date}."
        
        df_filtrado = df.query("sku == 'DVLAVADORAPREMIUM'")
        
        # Converter todas as colunas de data para datetime sem timezone
        for col in df_filtrado.columns:
            if pd.api.types.is_datetime64_any_dtype(df_filtrado[col]):
                df_filtrado[col] = df_filtrado[col].apply(lambda x: x.replace(tzinfo=None) if pd.notnull(x) else x)
            elif isinstance(df_filtrado[col].dtype, pd.DatetimeTZDtype):
                df_filtrado[col] = df_filtrado[col].dt.tz_localize(None)
        
        # Criar um novo workbook e selecionar a planilha ativa
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório por Modalidade"

        # Adicionar título
        ws['A1'] = f"Relatório por Modalidade - Período: {start_date} a {end_date}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:K1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Adicionar cabeçalhos
        headers = ['Data', 'Order ID', 'SKU', 'Nome do Produto', 'Status do Pedido', 'Status do Pagamento', 
                'Quantidade', 'Preço Unitário', 'Valor Pago (sem frete)', 'Custo Logístico', 'Tipo Logístico']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Adicionar dados
        for row_idx, row in enumerate(dataframe_to_rows(df_filtrado, index=False, header=False), start=3):
            for col_idx, value in enumerate(row, start=1):
                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in [8, 9, 10]:  # Colunas de valores monetários
                    cell.number_format = '"R$"#,##0.00'

        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar o arquivo
        file_name = f"Relatorio_Modalidade_{start_date.replace('/', '_')}_{end_date.replace('/', '_')}.xlsx"
        try:
            wb.save(file_name)
            return f"Relatório gerado com sucesso: {file_name}"
        except PermissionError:
            return f"Erro ao salvar o arquivo. Certifique-se de que {file_name} não está aberto em outro programa."
        

if __name__ == "__main__":
    try:
        ml_api = MercadoLivreAPI()
        
        #Exemplo de uso dos novos métodos
        start_date = "17/04/2025"
        end_date = "17/04/2025"

        #logger.info(f"Gerando relatório de vendas para o período de {start_date} a {end_date}")
        #salesDf = ml_api.get_sales_data(start_date,end_date)
        #salesDf.to_csv('meli_sales_data.csv')


        #logger.info(f"Gerando relatório geral para o período de {start_date} a {end_date}")
        #general_report = ml_api.generate_general_report(start_date, end_date)
        #print(general_report)
        
        #print("\n" + "="*60 + "\n")
        
        #logger.info(f"Gerando relatório por modalidade para o período de {start_date} a {end_date}")
        #modality_report = ml_api.generate_modality_report(start_date, end_date)
        #print(modality_report)

        #logger.info(f"Gerando relatório por modalidade para o período de {start_date} a {end_date}")
        #modality_report = ml_api.generate_modality_report_excel(start_date, end_date)
        #print(modality_report)        

                                 
        #Estoque        
        df = ml_api.gerar_relatorio_estoque()        
        if not df.empty:
            print("\nRelatório de Estoque Mercado Livre")
            print("=" * 60)
            print(df.to_string(index=False))
        else:
            print("Nenhum dado encontrado.")


        #Anúncios/ perguntas e Respostas
        #products = ml_api._get_active_items()

        #for product in products:
                        
        #    faqs = ml_api.get_all_questions_answers(product)
        #    print(f"Total de questões recuperadas do anúncio {product}: {len(faqs)}")





        #Perguntas
        #item_id = "MLB5213093324"
        #faqs = ml_api.get_all_questions_answers(item_id)

        #for faq in faqs:
        #    print(f"Pergunta: {faq['question_text']}")
        #    print(f"Resposta: {faq['answer_text']}")
        #    print("-" * 40)

        #print(f"Total de interações recuperadas: {len(faqs)}")
        

                    
    except MercadoLivreAPIError as e:
        print(f"Erro na execução: {str(e)}")